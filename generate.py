import pandas as pd
import matplotlib.pyplot as plt
import os
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Function to extract clean repository name
def clean_repo_name(repo_name):
    # Look for patterns like l3-angular-delta, l3-laravel-pharmalys, etc.
    patterns = [
        r'l\d+-(\w+)-([^_\s]+)',  # Matches l3-angular-delta or l3-laravel-pharmalys
        r'l\d+-(\w+)-([^_\s]+)-(\w+)'  # Matches l3-angular-delta-erp or l3-net-ipex-business
    ]
    
    for pattern in patterns:
        match = re.search(pattern, repo_name, re.IGNORECASE)
        if match:
            if len(match.groups()) == 2:
                tech, project = match.groups()
                return f"{tech}-{project}"
            elif len(match.groups()) == 3:
                tech, project, suffix = match.groups()
                return f"{tech}-{project}-{suffix}"
    
    # If no standard pattern, try to extract meaningful part
    if '_' in repo_name:
        # Try to extract from SELISEdigitalplatforms_l3-name_ID format
        parts = repo_name.split('_')
        if len(parts) >= 2:
            for part in parts:
                if part.startswith('l'):
                    return clean_repo_name(part)  # Recursively clean this part
    
    # If we can't extract a cleaner name, return the original
    return repo_name

# Function to filter data based on branch criteria
def filter_branch_data(df):
    # Remove any blank rows where any essential column is missing
    df = df.dropna(subset=['Repository Name', 'Branch'])
    
    branch_filters = ['stg', 'stage', 'stg-aks', 'stagging']
    mask = df['Branch'].str.contains('|'.join(branch_filters), case=False, na=False)
    return df[mask]

# Function to compare metrics and generate results
def compare_metrics(march_df, april_df, metric_name, min_diff=0):
    # Create a dictionary to store the results
    results = []
    
    # Create a unique identifier for each repo-branch combination
    march_df['RepoAndBranch'] = march_df['Repository Name'] + '___' + march_df['Branch']
    april_df['RepoAndBranch'] = april_df['Repository Name'] + '___' + april_df['Branch']
    
    # Get unique repo-branch combinations from both months
    march_repo_branches = set(march_df['RepoAndBranch'])
    april_repo_branches = set(april_df['RepoAndBranch'])
    
    # Find common repo-branch combinations
    common_repo_branches = march_repo_branches.intersection(april_repo_branches)
    
    # For each common repo-branch, compare the metric values
    for repo_branch in common_repo_branches:
        # Skip if the identifier is empty or NaN
        if pd.isna(repo_branch) or repo_branch == '':
            continue
        
        # Split the identifier back to repository and branch
        repo, branch = repo_branch.split('___')
        
        # Skip if repo is empty
        if pd.isna(repo) or repo == '':
            continue
            
        # Get the March value for this repo-branch, skip if missing
        march_row = march_df[march_df['RepoAndBranch'] == repo_branch]
        if march_row.empty or pd.isna(march_row[metric_name].values[0]):
            continue
        march_value = march_row[metric_name].values[0]
        
        # Get the April value for this repo-branch, skip if missing
        april_row = april_df[april_df['RepoAndBranch'] == repo_branch]
        if april_row.empty or pd.isna(april_row[metric_name].values[0]):
            continue
        april_value = april_row[metric_name].values[0]
        
        # Calculate the difference
        difference = april_value - march_value
        
        # For Code Smell, check if the absolute difference is >= 20 OR <= -20
        if metric_name == 'Code Smell' and abs(difference) < 20:
            continue
        
        # For other metrics, check if there's any change
        if metric_name != 'Code Smell' and difference == 0:
            continue
        
        # Get a clean repository name
        clean_name = clean_repo_name(repo)
        
        # Add to results
        results.append({
            'Repository Name': repo,
            'Branch': branch,
            'Clean Name': clean_name,
            f'{metric_name}_March': march_value,
            f'{metric_name}_April': april_value,
            f'{metric_name}_Difference': difference
        })
    
    # Convert results to DataFrame
    result_df = pd.DataFrame(results)
    
    return result_df

# Function to create Excel file with color coding
def create_excel_with_color(df, metric_name, output_file):
    # If no changes, create a simple excel with a message
    if df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{metric_name} Changes"
        ws.cell(row=1, column=1).value = f"No significant changes in {metric_name} between March and April"
        wb.save(output_file)
        print(f"No significant changes found for {metric_name}")
        return
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{metric_name} Changes"
    
    # Add headers and data to worksheet
    headers = [
        "Repository Name",
        "Branch", 
        "Clean Name",
        f"{metric_name} (March)", 
        f"{metric_name} (April)", 
        f"{metric_name} Difference"
    ]
    
    # Add the headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    # Add the data
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        ws.cell(row=row_num, column=1).value = row[0]  # Repository Name
        ws.cell(row=row_num, column=2).value = row[1]  # Branch
        ws.cell(row=row_num, column=3).value = row[2]  # Clean Name
        ws.cell(row=row_num, column=4).value = row[3]  # March value
        ws.cell(row=row_num, column=5).value = row[4]  # April value
        ws.cell(row=row_num, column=6).value = row[5]  # Difference
        
        # Apply color to the difference cell
        # Green if negative (improvement), Red if positive (regression)
        if row[5] < 0:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        else:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Create a vertical bar chart with positive and negative values going in opposite directions
    fig, ax = plt.figure(figsize=(10, 8)), plt.subplot(111)
    
    # Sort by difference for better visualization
    plot_df = df.sort_values(f'{metric_name}_Difference')
    
    # Add branch name to the clean name for the chart
    plot_df['Display_Name'] = plot_df['Clean Name'] + ' (' + plot_df['Branch'] + ')'
    
    # Plot the horizontal bars (which will appear as vertical when we flip the axes)
    # Separate positive and negative values
    pos_mask = plot_df[f'{metric_name}_Difference'] >= 0
    neg_mask = plot_df[f'{metric_name}_Difference'] < 0
    
    # SWAPPED: Plot positive values (regressions) on the LEFT
    if pos_mask.any():
        ax.barh(
            y=plot_df.loc[pos_mask, 'Display_Name'],
            width=-plot_df.loc[pos_mask, f'{metric_name}_Difference'],  # Negative width to go left
            color='red',
            label='Regression'
        )
    
    # SWAPPED: Plot negative values (improvements) on the RIGHT
    if neg_mask.any():
        ax.barh(
            y=plot_df.loc[neg_mask, 'Display_Name'],
            width=-plot_df.loc[neg_mask, f'{metric_name}_Difference'],  # Negative width to flip direction
            color='green',
            label='Improvement'
        )
    
    # Add a vertical line at x=0
    ax.axvline(0, color='black', linestyle='-', linewidth=0.5)
    
    plt.title(f'{metric_name} Difference (April - March)')
    plt.xlabel('Difference (absolute value)')
    plt.ylabel('Repository and Branch')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    plt.tight_layout()
    
    # Add legend
    if pos_mask.any() or neg_mask.any():
        plt.legend()
    
    # Save the chart
    chart_file = f"{metric_name}_chart.png"
    plt.savefig(chart_file)
    plt.close()
    
    # Add the chart to the Excel file
    img = Image(chart_file)
    img.width, img.height = 600, 400
    ws.add_image(img, 'H2')
    
    # Save the Excel workbook
    wb.save(output_file)
    
    # Clean up the temporary chart file
    os.remove(chart_file)

def main():
    try:
        # Load the Excel files (replace with your actual file paths)
        march_df = pd.read_excel('march_report.xlsx')
        april_df = pd.read_excel('april_report.xlsx')
        
        # Remove blank rows from both datasets by specifically checking essential columns
        # First, remove rows where Repository Name or Branch is missing
        march_df = march_df.dropna(subset=['Repository Name', 'Branch'])
        april_df = april_df.dropna(subset=['Repository Name', 'Branch'])
        
        # Also remove rows where the Repository Name is an empty string
        march_df = march_df[march_df['Repository Name'].str.strip() != '']
        april_df = april_df[april_df['Repository Name'].str.strip() != '']
        
        # Filter the data based on branch criteria
        march_filtered = filter_branch_data(march_df)
        april_filtered = filter_branch_data(april_df)
        
        # Compare and process each metric
        metrics = ['Code Smell', 'Duplications', 'Security Hotspot']
        
        for metric in metrics:
            # Only process repositories that have non-null values for this metric
            march_metric_filtered = march_filtered.dropna(subset=[metric])
            april_metric_filtered = april_filtered.dropna(subset=[metric])
            
            # Compare the metric between the two months
            result_df = compare_metrics(march_metric_filtered, april_metric_filtered, metric)
            
            # Create the output Excel file with color coding and chart
            output_file = f"{metric.replace(' ', '_')}_comparison.xlsx"
            create_excel_with_color(result_df, metric, output_file)
            
            if not result_df.empty:
                print(f"Generated {output_file} with {len(result_df)} repositories that had significant changes in {metric}")
                if metric == 'Code Smell':
                    print("Note: For Code Smell, only changes with absolute difference ≥ 20 are included")
        
        print("\nProcessing complete! All output files have been generated.")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()