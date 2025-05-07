import pandas as pd
import matplotlib.pyplot as plt
import os
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Function to extract clean repository name
def clean_repo_name(repo_name):
    # Look for patterns like l3-angular-delta, l3-laravel-pharmalys, etc.
    patterns = [
        r'l\d+-(\w+)-([^_\s]+)',  # Matches l3-angular-delta or l3-laravel-pharmalys
        r'l\d+-(\w+)-([^_\s]+)-(\w+)'  # Matches l3-angular-delta-erp or l3-net-ipex-business
    ]
    
    for pattern in patterns:
        match = re.search(pattern, repo_name, re.IGNORECASE)
        if match:
            if len(match.groups()) == 2:
                tech, project = match.groups()
                return f"{tech}-{project}"
            elif len(match.groups()) == 3:
                tech, project, suffix = match.groups()
                return f"{tech}-{project}-{suffix}"
    
    # If no standard pattern, try to extract meaningful part
    if '_' in repo_name:
        # Try to extract from SELISEdigitalplatforms_l3-name_ID format
        parts = repo_name.split('_')
        if len(parts) >= 2:
            for part in parts:
                if part.startswith('l'):
                    return clean_repo_name(part)  # Recursively clean this part
    
    # If we can't extract a cleaner name, return the original
    return repo_name

# Function to filter data based on branch criteria
def filter_branch_data(df):
    # Remove any blank rows where any essential column is missing
    df = df.dropna(subset=['Repository Name', 'Branch'])
    
    branch_filters = ['stg', 'stage', 'stg-aks', 'stagging']
    mask = df['Branch'].str.contains('|'.join(branch_filters), case=False, na=False)
    return df[mask]

# Function to compare metrics and generate results
def compare_metrics(first_month, second_month, metric_name, min_diff=0):
    # Create a dictionary to store the results
    results = []
    
    # Create a unique identifier for each repo-branch combination
    first_month['RepoAndBranch'] = first_month['Repository Name'] + '___' + first_month['Branch']
    second_month['RepoAndBranch'] = second_month['Repository Name'] + '___' + second_month['Branch']
    
    # Get unique repo-branch combinations from both months
    first_repo_branches = set(first_month['RepoAndBranch'])
    second_repo_branches = set(second_month['RepoAndBranch'])
    
    # Find common repo-branch combinations
    common_repo_branches = first_repo_branches.intersection(second_repo_branches)
    
    # For each common repo-branch, compare the metric values
    for repo_branch in common_repo_branches:
        # Skip if the identifier is empty or NaN
        if pd.isna(repo_branch) or repo_branch == '':
            continue
        
        # Split the identifier back to repository and branch
        repo, branch = repo_branch.split('___')
        
        # Skip if repo is empty
        if pd.isna(repo) or repo == '':
            continue
            
        # Get the first value for this repo-branch, skip if missing
        first_row = first_month[first_month['RepoAndBranch'] == repo_branch]
        if first_row.empty or pd.isna(first_row[metric_name].values[0]):
            continue
        first_value = first_row[metric_name].values[0]
        
        # Get the second value for this repo-branch, skip if missing
        second_row = second_month[second_month['RepoAndBranch'] == repo_branch]
        if second_row.empty or pd.isna(second_row[metric_name].values[0]):
            continue
        second_value = second_row[metric_name].values[0]
        
        # Calculate the difference
        difference = second_value - first_value
        
        # For Code Smell, check if the absolute difference is >= 20 OR <= -20
        if metric_name == 'Code Smell' and abs(difference) < 20:
            continue
        
        # For other metrics, check if there's any change
        if metric_name != 'Code Smell' and difference == 0:
            continue
        
        # Get a clean repository name
        clean_name = clean_repo_name(repo)
        
        # Add to results
        results.append({
            'Repository Name': repo,
            'Branch': branch,
            'Clean Name': clean_name,
            f'{metric_name}_first': first_value,
            f'{metric_name}_second': second_value,
            f'{metric_name}_Difference': difference,
            'Metric': metric_name  # Add the metric name for later identification
        })
    
    # Convert results to DataFrame
    result_df = pd.DataFrame(results)
    
    return result_df

# Function to create Excel file with color coding
def create_excel_with_color(df, metric_name, output_file):
    # If no changes, create a simple excel with a message
    if df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{metric_name} Changes"
        ws.cell(row=1, column=1).value = f"No significant changes in {metric_name} between first and second"
        wb.save(output_file)
        print(f"No significant changes found for {metric_name}")
        return
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{metric_name} Changes"
    
    # Add headers and data to worksheet
    headers = [
        "Repository Name",
        "Branch", 
        "Clean Name",
        f"{metric_name} (first)", 
        f"{metric_name} (second)", 
        f"{metric_name} Difference"
    ]
    
    # Add the headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    # Add the data
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        ws.cell(row=row_num, column=1).value = row[0]  # Repository Name
        ws.cell(row=row_num, column=2).value = row[1]  # Branch
        ws.cell(row=row_num, column=3).value = row[2]  # Clean Name
        ws.cell(row=row_num, column=4).value = row[3]  # first value
        ws.cell(row=row_num, column=5).value = row[4]  # second value
        ws.cell(row=row_num, column=6).value = row[5]  # Difference
        
        # Apply color to the difference cell
        # Green if negative (improvement), Red if positive (regression)
        if row[5] < 0:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        else:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Create a vertical bar chart with positive and negative values going in opposite directions
    fig, ax = plt.figure(figsize=(10, 8)), plt.subplot(111)
    
    # Sort by difference for better visualization
    plot_df = df.sort_values(f'{metric_name}_Difference')
    
    # Add branch name to the clean name for the chart
    plot_df['Display_Name'] = plot_df['Clean Name'] + ' (' + plot_df['Branch'] + ')'
    
    # Plot the horizontal bars (which will appear as vertical when we flip the axes)
    # Separate positive and negative values
    pos_mask = plot_df[f'{metric_name}_Difference'] >= 0
    neg_mask = plot_df[f'{metric_name}_Difference'] < 0
    
    # SWAPPED: Plot positive values (regressions) on the LEFT
    if pos_mask.any():
        ax.barh(
            y=plot_df.loc[pos_mask, 'Display_Name'],
            width=-plot_df.loc[pos_mask, f'{metric_name}_Difference'],  # Negative width to go left
            color='red',
            label='Regression'
        )
    
    # SWAPPED: Plot negative values (improvements) on the RIGHT
    if neg_mask.any():
        ax.barh(
            y=plot_df.loc[neg_mask, 'Display_Name'],
            width=-plot_df.loc[neg_mask, f'{metric_name}_Difference'],  # Negative width to flip direction
            color='green',
            label='Improvement'
        )
    
    # Add a vertical line at x=0
    ax.axvline(0, color='black', linestyle='-', linewidth=0.5)
    
    plt.title(f'{metric_name} Difference (May - April)')
    plt.xlabel('Difference (absolute value)')
    plt.ylabel('Repository and Branch')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    plt.tight_layout()
    
    # Add legend
    if pos_mask.any() or neg_mask.any():
        plt.legend()
    
    # Save the chart
    chart_file = f"{metric_name}_chart.png"
    plt.savefig(chart_file)
    plt.close()
    
    # Add the chart to the Excel file
    img = Image(chart_file)
    img.width, img.height = 600, 400
    ws.add_image(img, 'H2')
    
    # Save the Excel workbook
    wb.save(output_file)
    
    # Clean up the temporary chart file
    os.remove(chart_file)

# IMPROVED: Create a compact summary chart with boxed/separated sections for each metric
def create_summary_chart(all_results, top_n=3):
    # Create a figure with subplots for each metric
    num_metrics = len(all_results)
    if num_metrics == 0:
        print("No metrics with significant changes to summarize")
        return None
    
    # Filter out metrics with no improvements
    metrics_with_improvements = {}
    for metric_name, df in all_results.items():
        improvements = df[df[f'{metric_name}_Difference'] < 0].copy()
        if not improvements.empty:
            metrics_with_improvements[metric_name] = df
    
    if not metrics_with_improvements:
        print("No improvements found in any metric")
        return None
    
    # Update num_metrics to only count those with improvements
    num_metrics = len(metrics_with_improvements)
    
    # Use a professional color palette
    bar_color = '#2ca02c'  # Green for improvements
    
    # Set figure style for professional presentation
    plt.style.use('default')
    
    # Calculate optimal figure dimensions based on data
    # Find the longest repository name to adjust width
    max_name_length = 0
    for metric_name, df in metrics_with_improvements.items():
        improvements = df[df[f'{metric_name}_Difference'] < 0].copy().head(top_n)
        if not improvements.empty:
            for name in improvements['Clean Name']:
                max_name_length = max(max_name_length, len(str(name)))
    
    # Adjust width based on maximum name length
    base_width = 8
    width_adjustment = min(max_name_length * 0.1, 3)  # Cap the width adjustment
    fig_width = base_width + width_adjustment
    
    # Height based on number of metrics and bars per metric + extra space for boxes
    fig_height = 1.8 * top_n * num_metrics
    
    # Create figure - use constrained_layout for better spacing with boxes
    fig = plt.figure(figsize=(fig_width, fig_height), constrained_layout=False)
    
    # Create a grid for the subplots with larger gaps between metric sections
    grid_spec = fig.add_gridspec(nrows=num_metrics, ncols=1, hspace=0.6, 
                                 left=0.25, right=0.85, top=0.95, bottom=0.05)
    
    # Add a title to the entire figure
    plt.suptitle('Top Quality Improvements (May vs April)', 
                fontsize=14, y=0.98, fontweight='bold')
    
    # Process each metric
    for i, (metric_name, df) in enumerate(metrics_with_improvements.items()):
        # Create subplot
        ax = fig.add_subplot(grid_spec[i, 0])
        
        # Filter for improvements only (negative differences)
        improvements = df[df[f'{metric_name}_Difference'] < 0].copy()
        
        # Sort by difference (most negative = biggest improvement)
        improvements = improvements.sort_values(by=f'{metric_name}_Difference')
        
        # Get top N improvements
        top_improvements = improvements.head(top_n)
        
        # Create shorter display names
        top_improvements['Display_Name'] = top_improvements['Clean Name'] + ' (' + top_improvements['Branch'].str.split('-').str[0] + ')'
        
        # Plot horizontal bars
        bars = ax.barh(
            y=top_improvements['Display_Name'],
            width=-top_improvements[f'{metric_name}_Difference'],  # Make positive for visualization
            color=bar_color,
            alpha=0.9,
            height=0.6,  # Slightly thicker bars for better visibility
            edgecolor='none'
        )
        
        # Set tight x-axis limits to avoid wasted space
        # Find the maximum bar width and add a small margin
        max_bar_width = -top_improvements[f'{metric_name}_Difference'].min() * 1.15
        ax.set_xlim(0, max_bar_width)
        
        # Customize the plot
        ax.set_title(f'Top {metric_name} Improvements', fontsize=12, fontweight='bold', pad=8)
        
        # Create a box around each section by adding visible spines
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1)
            spine.set_edgecolor('gray')
            spine.set_linestyle('-')
        
        # Custom background for better separation
        ax.set_facecolor('#f8f8f8')  # Light gray background
        
        # Custom x-axis with minimal ticks
        if max_bar_width > 0:
            ax.set_xticks([0, max_bar_width/2, max_bar_width])
            ax.set_xticklabels(['0', f'{int(max_bar_width/2)}', f'{int(max_bar_width)}'])
        
        # Add value labels directly on bars
        for j, bar in enumerate(bars):
            diff_value = abs(top_improvements.iloc[j][f'{metric_name}_Difference'])
            before_value = int(top_improvements.iloc[j][f'{metric_name}_first'])
            after_value = int(top_improvements.iloc[j][f'{metric_name}_second'])
            
            # Add improvement value at end of bar
            ax.text(
                diff_value + (max_bar_width * 0.02),  # Small offset
                j,
                f"-{int(diff_value)}",
                va='center',
                fontsize=10,
                fontweight='bold',
                color='darkgreen'
            )
            
            # Add from→to at the middle of the bar
            bar_width = bar.get_width()
            if bar_width > 20:  # Only add if there's enough space
                ax.text(
                    bar_width/2.5,  # Position proportionally inside bar
                    j,
                    f"{before_value}→{after_value}",
                    va='center',
                    ha='center',
                    fontsize=9,
                    color='black'
                )
    
    # Adjust layout to ensure proper spacing
    plt.tight_layout(rect=[0, 0, 1, 0.97], h_pad=3.0)
    
    # Save the summary chart with high resolution and minimal padding
    summary_file = "top_improvements_summary.png"
    plt.savefig(summary_file, dpi=300, bbox_inches='tight', pad_inches=0.3)
    plt.close()
    
    print(f"Generated compact presentation-ready chart with boxed sections: {summary_file}")
    
    return summary_file

# New function to create a summary Excel with only data (no chart)
def create_summary_excel(all_results):
    # Create a summary Excel file with the top improvements data
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Improvements Summary"
    
    # Add a title
    ws.cell(row=1, column=1).value = "Summary of Top Improvements Across All Metrics (May - April)"
    
    # Add details for each metric
    current_row = 3
    
    for metric, df in all_results.items():
        # Filter for improvements only (negative differences)
        improvements = df[df[f'{metric}_Difference'] < 0].copy()
        
        if improvements.empty:
            ws.cell(row=current_row, column=1).value = f"No improvements found for {metric}"
            current_row += 2
            continue
        
        # Sort by difference (most negative = biggest improvement)
        improvements = improvements.sort_values(by=f'{metric}_Difference')
        
        # Get top improvements
        top_improvements = improvements.head(5)  # Show up to 5 in the spreadsheet for detail
        
        # Add header for this metric
        ws.cell(row=current_row, column=1).value = f"Top Improvements - {metric}"
        current_row += 1
        
        # Add column headers
        headers = ["Repository", "Branch", "First Value", "Second Value", "Difference"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col_num).value = header
        current_row += 1
        
        # Add data rows
        for _, row in top_improvements.iterrows():
            ws.cell(row=current_row, column=1).value = row['Clean Name']
            ws.cell(row=current_row, column=2).value = row['Branch']
            ws.cell(row=current_row, column=3).value = row[f'{metric}_first']
            ws.cell(row=current_row, column=4).value = row[f'{metric}_second']
            ws.cell(row=current_row, column=5).value = row[f'{metric}_Difference']
            
            # Color the difference cell green (improvement)
            ws.cell(row=current_row, column=5).fill = PatternFill(
                start_color="00FF00", end_color="00FF00", fill_type="solid"
            )
            
            current_row += 1
        
        # Add spacing between metrics
        current_row += 3
    
    # Save the summary Excel file
    summary_excel = "top_improvements_summary.xlsx"
    wb.save(summary_excel)
    print(f"Generated summary Excel file: {summary_excel}")
    
    return summary_excel

def main():
    try:
        # Load the Excel files (replace with your actual file paths)
        first_month = pd.read_excel('april_report.xlsx')
        second_month = pd.read_excel('may_report.xlsx')
        
        # Remove blank rows from both datasets by specifically checking essential columns
        # First, remove rows where Repository Name or Branch is missing
        first_month = first_month.dropna(subset=['Repository Name', 'Branch'])
        second_month = second_month.dropna(subset=['Repository Name', 'Branch'])
        
        # Also remove rows where the Repository Name is an empty string
        first_month = first_month[first_month['Repository Name'].str.strip() != '']
        second_month = second_month[second_month['Repository Name'].str.strip() != '']
        
        # Filter the data based on branch criteria
        first_filtered = filter_branch_data(first_month)
        second_filtered = filter_branch_data(second_month)
        
        # Compare and process each metric
        metrics = ['Code Smell', 'Duplications', 'Security Hotspot']
        all_results = {}  # Store results for all metrics for the summary chart
        
        for metric in metrics:
            # Only process repositories that have non-null values for this metric
            first_metric_filtered = first_filtered.dropna(subset=[metric])
            second_metric_filtered = second_filtered.dropna(subset=[metric])
            
            # Compare the metric between the two months
            result_df = compare_metrics(first_metric_filtered, second_metric_filtered, metric)
            
            # Store the results for summary chart
            if not result_df.empty:
                all_results[metric] = result_df
            
            # Create the output Excel file with color coding and chart
            output_file = f"{metric.replace(' ', '_')}_comparison.xlsx"
            create_excel_with_color(result_df, metric, output_file)
            
            if not result_df.empty:
                print(f"Generated {output_file} with {len(result_df)} repositories that had significant changes in {metric}")
                if metric == 'Code Smell':
                    print("Note: For Code Smell, only changes with absolute difference ≥ 20 are included")
        
        # Handle the summary outputs as separate files
        if all_results:
            # 1. Create summary chart as separate image file
            summary_image = create_summary_chart(all_results, top_n=3)
            
            # 2. Create a separate summary Excel file with just the data (no chart)
            summary_excel = create_summary_excel(all_results)
            
            print("\nSummary files generated:")
            print(f"- Image: {summary_image}")
            print(f"- Excel: {summary_excel}")
        
        print("\nProcessing complete! All output files have been generated.")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()